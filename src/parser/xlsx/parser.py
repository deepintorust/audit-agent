from __future__ import annotations

from openpyxl import load_workbook

from src.parser.base import Parser
from src.parser.ir import Block, DocumentIR


class XlsxParser(Parser):
    def parse(self, path: str) -> DocumentIR:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        blocks: list[Block] = []
        for sheet in wb.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                continue
            headers = [str(x).strip() if x is not None else "" for x in header_row]
            for idx, row in enumerate(rows, start=2):
                kv = {}
                for h, v in zip(headers, row, strict=False):
                    if not h:
                        continue
                    if v is None:
                        continue
                    kv[h] = str(v)
                if not kv:
                    continue
                text = "；".join([f"{k}:{v}" for k, v in kv.items()])
                blocks.append(
                    Block(kind="table_row", text=text, meta={"sheet": sheet.title, "row": idx, "headers": headers})
                )
        return DocumentIR(blocks=blocks)

